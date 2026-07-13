"""XLSX-Handler (openpyxl) mit Formel-Token-AST.

Strategie:

* Jede Zelle wird einzeln transformiert (String-Zellen).
* Numerische Zellen bleiben unangetastet (Phase 1 — siehe Megaprompt §5.4).
* Formeln (``cell.data_type == 'f'``) werden via
  :class:`openpyxl.formula.tokenizer.Tokenizer` zerlegt. Nur Token vom
  Subtyp ``TEXT`` werden transformiert; Bereiche, Funktionsnamen, Sheet-
  Namen und Zellbezüge bleiben strukturell unverändert.
* Excel-escaped Quotes (`""` innerhalb von Strings) werden korrekt
  behandelt — der Tokenizer übernimmt das Roundtripping.
* Sheet-Reihenfolge, Named Ranges und Sheet-übergreifende Referenzen
  (z. B. ``='Mandant Hofer'!A1``) bleiben erhalten.
* **Named Ranges**: Die Workbook-DefinedNames werden ebenfalls
  durchgegangen — String-Literale in Named-Range-Formeln werden
  transformiert. Die Range-Namen selbst (``MWST_SATZ``) bleiben
  unverändert, weil sie keine PII enthalten.
"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl.formula.tokenizer import Tokenizer

from pseudokrat.formats.base import (
    FormatProcessResult,
    TextTransform,
    derive_default_output,
)

if TYPE_CHECKING:  # pragma: no cover
    pass


def _unescape_excel_string(token_value: str) -> str:
    """Excel-String-Literal → Python-String.

    Token-Werte vom Tokenizer kommen MIT den umschließenden Quotes.
    Doppelte Quotes inside (`""`) sind Excel-Escape für ein einfaches `"`.
    """
    if not (token_value.startswith('"') and token_value.endswith('"')):
        return token_value
    inner = token_value[1:-1]
    return inner.replace('""', '"')


def _escape_excel_string(value: str) -> str:
    """Python-String → Excel-String-Literal, inkl. Quote-Escaping."""
    return '"' + value.replace('"', '""') + '"'


def _transform_formula(formula: str, transform: TextTransform) -> str:
    """Zerlege die Formel mit dem openpyxl-Tokenizer und transformiere
    ausschließlich TEXT-Operanden.

    Erhält Whitespace, Operator-Symbole, Funktionsnamen, Range-Operatoren
    und Sheet-Verweise bit-exakt.
    """
    if not formula.startswith("="):
        # Fallback: openpyxl-Tokenizer erwartet eine vollständige Formel
        # mit führendem '=' — kommt vor bei manuell konstruierten Formel-
        # Strings in seltenen Edge-Fällen.
        formula = "=" + formula
        strip_lead = True
    else:
        strip_lead = False

    tokens = Tokenizer(formula).items
    rebuilt: list[str] = []
    for tok in tokens:
        if tok.type == "OPERAND" and tok.subtype == "TEXT":
            inner = _unescape_excel_string(tok.value)
            rebuilt.append(_escape_excel_string(transform(inner)))
        elif tok.type == "WHITE-SPACE":
            rebuilt.append(tok.value)
        else:
            rebuilt.append(tok.value)
    result = "=" + "".join(rebuilt)
    return result[1:] if strip_lead else result


def _transform_headers_footers(sheet: object, transform: TextTransform) -> int:
    """Anonymisiert Kopf-/Fusszeilen (links/mitte/rechts) eines Arbeitsblatts.

    Kopf-/Fusszeilen tragen häufig Briefkopf, Mandant, „Vertraulich – <Name>"
    o. ä. — ein Kanal, den die reine Zell-Iteration nicht erfasst.
    """
    changed = 0
    for hf_name in (
        "oddHeader", "oddFooter", "evenHeader", "evenFooter",
        "firstHeader", "firstFooter",
    ):
        hf = getattr(sheet, hf_name, None)
        if hf is None:
            continue
        for part_name in ("left", "center", "right"):
            part = getattr(hf, part_name, None)
            text = getattr(part, "text", None)
            if text:
                new_text = transform(text)
                if new_text != text:
                    part.text = new_text
                    changed += 1
    return changed


class XlsxHandler:
    name = "xlsx"
    suffixes: tuple[str, ...] = (".xlsx",)

    def supports(self, path: Path) -> bool:
        return path.suffix.lower() in self.suffixes

    def default_output_path(self, input_path: Path, suffix: str = "anon") -> Path:
        return derive_default_output(input_path, suffix=suffix)

    @staticmethod
    def _transform_formula(formula: str, transform: TextTransform) -> str:
        """Public-Helper für Tests — delegiert auf Modul-Funktion."""
        return _transform_formula(formula, transform)

    def process(
        self,
        input_path: Path,
        output_path: Path,
        transform: TextTransform,
        *,
        permute_numeric_columns_with: object | None = None,
    ) -> FormatProcessResult:
        """Anonymisiere eine XLSX-Datei.

        ``permute_numeric_columns_with`` ist optional ein
        :class:`pseudokrat.dp.PermutationKey`. Wenn gesetzt, werden alle
        numerischen Spalten **mit Ausnahme von Zeile 1** rangbewahrend
        permutiert (Mittelwert/Summe bleiben, Zuordnung Mandant→Betrag
        verschwindet).
        """
        from openpyxl import load_workbook

        workbook = load_workbook(str(input_path))
        processed = 0
        skipped = 0
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    # Zell-Kommentare/Notizen (verstecktes PII-Leck).
                    comment = cell.comment
                    if comment is not None and comment.text:
                        new_comment = transform(comment.text)
                        if new_comment != comment.text:
                            comment.text = new_comment
                            processed += 1
                    value = cell.value
                    if value is None:
                        continue
                    if cell.data_type == "f" and isinstance(value, str):
                        new_value = self._transform_formula(value, transform)
                        if new_value != value:
                            cell.value = new_value
                            processed += 1
                        else:
                            skipped += 1
                        continue
                    if isinstance(value, str):
                        new_value = transform(value)
                        if new_value != value:
                            cell.value = new_value
                            processed += 1
                        else:
                            skipped += 1
                    else:
                        skipped += 1
            # Kopf-/Fusszeilen (Briefkopf/Mandant landet oft hier).
            processed += _transform_headers_footers(sheet, transform)

        if permute_numeric_columns_with is not None:
            from pseudokrat.dp.numeric_permute import (
                PermutationKey,
                shuffle_numeric_columns,
            )

            if isinstance(permute_numeric_columns_with, PermutationKey):
                permuted = shuffle_numeric_columns(workbook, permute_numeric_columns_with)
                processed += permuted

        # Named Ranges: deren formula-attribut kann String-Literale enthalten,
        # die zur Mandantenliste gehören. Wir transformieren analog zu
        # Zellformeln; reine A1-Referenzen bleiben unverändert.
        defined_names = getattr(workbook, "defined_names", None)
        if defined_names is not None:
            # openpyxl ≥ 3.1: workbook.defined_names ist ein dict-ähnliches Objekt.
            for name in list(defined_names):
                dn = defined_names[name]
                attr = getattr(dn, "attr_text", None) or getattr(dn, "value", None)
                if isinstance(attr, str) and '"' in attr:
                    new_attr = self._transform_formula(attr, transform)
                    if new_attr != attr:
                        if hasattr(dn, "attr_text"):
                            dn.attr_text = new_attr
                        else:
                            dn.value = new_attr
                        processed += 1

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(str(output_path))
        return FormatProcessResult(
            input_path=input_path,
            output_path=output_path,
            segments_processed=processed,
            segments_skipped=skipped,
        )

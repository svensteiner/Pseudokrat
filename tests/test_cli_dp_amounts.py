"""CLI-Test für `pseudokrat anonymize --dp-amounts ...xlsx`."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from pseudokrat.cli import main


@pytest.fixture(autouse=True)
def _env(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("PSEUDOKRAT_DATA_DIR", str(tmp_path / "data"))
    monkeypatch.setenv("PSEUDOKRAT_DISABLE_ML", "1")


def _make_workbook(path: Path) -> list[int]:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Mandant"
    ws["B1"] = "Betrag"
    amounts = [110, 220, 330, 440, 550, 660, 770]
    for r, (m, b) in enumerate(zip("ABCDEFG", amounts, strict=True), start=2):
        ws.cell(row=r, column=1, value=f"Mandant {m}")
        ws.cell(row=r, column=2, value=b)
    wb.save(path)
    return amounts


def test_cli_dp_amounts_preserves_sum(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    src = tmp_path / "saldenliste.xlsx"
    out = tmp_path / "saldenliste.anon.xlsx"
    amounts = _make_workbook(src)

    rc = main(
        [
            "anonymize",
            "--profile",
            "default",
            "--password",
            "supergeheim",
            "--input",
            str(src),
            "--output",
            str(out),
            "--dp-amounts",
        ]
    )
    err = capsys.readouterr().err
    assert rc == 0, err
    assert out.exists()

    wb = load_workbook(out)
    ws = wb.active
    new_amounts = [ws.cell(row=r, column=2).value for r in range(2, 9)]
    assert sum(new_amounts) == sum(amounts)
    assert sorted(new_amounts) == sorted(amounts)


def test_cli_dp_amounts_warning_for_non_xlsx(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    src = tmp_path / "memo.txt"
    src.write_text("Hofer Bau GmbH zahlt.")
    out = tmp_path / "memo.anon.txt"

    rc = main(
        [
            "anonymize",
            "--profile",
            "default",
            "--password",
            "supergeheim",
            "--input",
            str(src),
            "--output",
            str(out),
            "--dp-amounts",
        ]
    )
    err = capsys.readouterr().err
    assert rc == 0
    assert "--dp-amounts" in err

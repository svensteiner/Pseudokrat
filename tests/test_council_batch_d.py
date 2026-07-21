"""Tests fuer Council-Batch-D: XLSX-Kommentare/Kopfzeilen (#10),
PDF-Annotationen (#13), Rest-PII-Gate-Textextraktion (#2)."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment

from pseudokrat import watcher
from pseudokrat.formats.xlsx_handler import XlsxHandler


def _mask_hankook(text: str) -> str:
    return text.replace("Hankook", "<COMPANY_001>")


class TestXlsxHiddenChannels:
    def test_comment_and_header_anonymized(self, tmp_path: Path) -> None:
        src = tmp_path / "in.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Zahl"
        ws["A1"].comment = Comment("Notiz zu Hankook", "tester")
        ws.oddHeader.center.text = "Hankook Bilanz"
        ws.oddFooter.right.text = "Vertraulich Hankook"
        wb.save(src)

        out = tmp_path / "out.xlsx"
        XlsxHandler().process(src, out, transform=_mask_hankook)

        wb2 = load_workbook(out)
        ws2 = wb2.active
        assert "Hankook" not in (ws2["A1"].comment.text or "")
        assert "Hankook" not in (ws2.oddHeader.center.text or "")
        assert "Hankook" not in (ws2.oddFooter.right.text or "")


class TestGateExtraction:
    def test_office_xml_text_extracted(self, tmp_path: Path) -> None:
        src = tmp_path / "g.xlsx"
        wb = Workbook()
        wb.active["A1"] = "AT611904300234573201"
        wb.save(src)
        text = watcher.extract_text_for_gate(src)
        assert "AT611904300234573201" in text

    def test_plain_text_extracted(self, tmp_path: Path) -> None:
        src = tmp_path / "a.txt"
        src.write_text("hallo welt", encoding="utf-8")
        assert "hallo welt" in watcher.extract_text_for_gate(src)


class TestPdfAnnotationsRemoved:
    def test_annotations_stripped(self, tmp_path: Path, store_and_audit) -> None:  # noqa: ANN001
        pymupdf = pytest.importorskip("pymupdf")
        from pseudokrat.anonymizer import Anonymizer
        from pseudokrat.recognizers import recognizers_for_store

        store, audit = store_and_audit
        doc = pymupdf.open()
        page = doc.new_page()
        page.insert_text((72, 72), "Bericht ohne kritischen Text.")
        page.add_text_annot((200, 200), "Kommentar von Sven Steiner")
        src = tmp_path / "a.pdf"
        doc.save(str(src))
        doc.close()

        anon = Anonymizer(
            store=store,
            recognizers=recognizers_for_store(store),
            detector=None,
            audit_log=audit,
            model_version="disabled",
        )
        out = tmp_path / "o.pdf"
        watcher.redact_pdf(src, out, anon, store, remove_logos=False, ocr=None, log=lambda _m: None)

        d2 = pymupdf.open(str(out))
        remaining = [a for p in d2 for a in list(p.annots() or [])]
        assert remaining == []
        d2.close()

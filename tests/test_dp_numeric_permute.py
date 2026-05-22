"""Tests für die rangbewahrende Permutation von numerischen Spalten."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from pseudokrat.dp.numeric_permute import (
    PermutationKey,
    column_stats,
    permutation_key_from_secret,
    shuffle_column,
    shuffle_numeric_columns,
)
from pseudokrat.formats.xlsx_handler import XlsxHandler


def test_shuffle_column_preserves_multiset() -> None:
    values = [100, 200, 300, 400, 500, 600, 700, 800, 900, 1000]
    result = shuffle_column(values, seed=42)
    assert sorted(v for v in result if v is not None) == sorted(values)


def test_shuffle_column_preserves_sum_and_mean() -> None:
    values = [100, 250, 333, 444, 555, 678, 911]
    result = shuffle_column(values, seed=12345)
    assert column_stats([v for v in result if v is not None]) == column_stats(values)


def test_shuffle_column_actually_shuffles_for_large_input() -> None:
    """Bei 100 Werten ist die Wahrscheinlichkeit, dass NICHT mind. ein Wert
    seinen Platz wechselt, vernachlässigbar (≈ 1/100!).
    """
    values = list(range(100))
    result = shuffle_column(values, seed=99)
    diff_positions = sum(1 for a, b in zip(values, result, strict=True) if a != b)
    assert diff_positions > 50


def test_shuffle_column_handles_none_holes() -> None:
    values = [10, None, 20, None, 30]
    result = shuffle_column(values, seed=1)
    # None-Positionen bleiben None.
    assert result[1] is None
    assert result[3] is None
    # Die nicht-None-Werte sind dieselben (Multiset).
    non_none = sorted(v for v in result if v is not None)
    assert non_none == [10, 20, 30]


def test_shuffle_is_deterministic_for_same_seed() -> None:
    values = [1, 2, 3, 4, 5, 6, 7]
    a = shuffle_column(values, seed=777)
    b = shuffle_column(values, seed=777)
    assert a == b


def test_shuffle_differs_for_different_seeds() -> None:
    values = list(range(50))
    a = shuffle_column(values, seed=1)
    b = shuffle_column(values, seed=2)
    assert a != b
    assert sorted(v for v in a if v is not None) == sorted(
        v for v in b if v is not None
    )


def test_permutation_key_seed_for_is_stable() -> None:
    key = PermutationKey(key_bytes=b"\x42" * 32)
    s1 = key.seed_for("Sheet1", "B")
    s2 = key.seed_for("Sheet1", "B")
    s3 = key.seed_for("Sheet1", "C")
    assert s1 == s2
    assert s1 != s3


def test_permutation_key_from_secret_is_domain_separated() -> None:
    """Verschiedene Sub-Domain-Keys müssen sich unterscheiden."""
    secret = b"masterkey-bytes-32-abcdefghijklmn"
    k1 = permutation_key_from_secret(secret)
    k2 = permutation_key_from_secret(b"other-secret")
    assert k1.key_bytes != k2.key_bytes
    assert len(k1.key_bytes) == 32


def test_shuffle_numeric_columns_skips_header_row(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Mandant"
    ws["B1"] = "Betrag"
    for r, (mandant, betrag) in enumerate(
        [("M-001", 100), ("M-002", 200), ("M-003", 300), ("M-004", 400)],
        start=2,
    ):
        ws.cell(row=r, column=1, value=mandant)
        ws.cell(row=r, column=2, value=betrag)

    key = PermutationKey(key_bytes=b"\x01" * 32)
    count = shuffle_numeric_columns(wb, key)
    assert count == 1  # nur Spalte B wurde permutiert (Spalte A ist String)

    # Header bleibt unangetastet:
    assert ws["A1"].value == "Mandant"
    assert ws["B1"].value == "Betrag"

    # Werte sind shuffled, aber Multiset stimmt:
    new_amounts = [ws.cell(row=r, column=2).value for r in range(2, 6)]
    assert sorted(new_amounts) == [100, 200, 300, 400]


def test_xlsx_handler_with_dp_amounts_preserves_sum(tmp_path: Path) -> None:
    """E2E: XlsxHandler.process mit Permutations-Key liefert dieselbe Summe."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Mandant"
    ws["B1"] = "Betrag"
    original_amounts = [100, 250, 333, 444, 555]
    for r, (mandant, betrag) in enumerate(
        zip(["M-001", "M-002", "M-003", "M-004", "M-005"], original_amounts, strict=True),
        start=2,
    ):
        ws.cell(row=r, column=1, value=mandant)
        ws.cell(row=r, column=2, value=betrag)

    input_path = tmp_path / "in.xlsx"
    output_path = tmp_path / "out.xlsx"
    wb.save(input_path)

    key = PermutationKey(key_bytes=b"\x99" * 32)
    handler = XlsxHandler()
    handler.process(
        input_path,
        output_path,
        transform=lambda s: f"<{s}>",
        permute_numeric_columns_with=key,
    )

    wb2 = load_workbook(output_path)
    ws2 = wb2.active
    new_amounts = [ws2.cell(row=r, column=2).value for r in range(2, 7)]
    assert sum(new_amounts) == sum(original_amounts)
    assert sorted(new_amounts) == sorted(original_amounts)


def test_xlsx_handler_without_key_does_not_permute(tmp_path: Path) -> None:
    """Ohne Permutations-Key bleiben Beträge an ihrem Platz."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Mandant"
    ws["B1"] = "Betrag"
    for r, (m, b) in enumerate(
        [("M-001", 100), ("M-002", 200), ("M-003", 300)], start=2
    ):
        ws.cell(row=r, column=1, value=m)
        ws.cell(row=r, column=2, value=b)
    input_path = tmp_path / "in.xlsx"
    output_path = tmp_path / "out.xlsx"
    wb.save(input_path)

    XlsxHandler().process(input_path, output_path, transform=lambda s: f"<{s}>")
    wb2 = load_workbook(output_path)
    assert wb2.active["B2"].value == 100
    assert wb2.active["B3"].value == 200
    assert wb2.active["B4"].value == 300


def test_shuffle_column_single_value_unchanged() -> None:
    assert shuffle_column([42], seed=0) == [42]


def test_shuffle_column_all_none_unchanged() -> None:
    assert shuffle_column([None, None, None], seed=0) == [None, None, None]


def test_column_stats_empty() -> None:
    s = column_stats([])
    assert s["count"] == 0
    assert s["mean"] == 0.0

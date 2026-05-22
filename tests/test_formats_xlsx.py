"""Tests für XLSX-Handler — Phase-4-Vorzieher."""

from __future__ import annotations

from pathlib import Path

from pseudokrat.formats.xlsx_handler import XlsxHandler


def _build_xlsx(path: Path) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Salden"
    ws.append(["Mandant", "Saldo"])
    ws.append(["Hofer Bau GmbH", 1000])
    ws.append(["Hofer Bau GmbH", 2000])
    ws.append(["Müller AG", 5000])
    # Summe-Formel mit Mandantenreferenz
    ws["C2"] = '=SUMIF(A:A,"Hofer Bau GmbH",B:B)'
    ws["C3"] = "=SUM(B2:B4)"  # ohne String-Literal — bleibt unverändert
    wb.save(str(path))


def test_xlsx_string_cells_consistent(tmp_path: Path) -> None:
    inp = tmp_path / "salden.xlsx"
    out = tmp_path / "salden.anon.xlsx"
    _build_xlsx(inp)

    mapping: dict[str, str] = {}

    def transform(text: str) -> str:
        if text in {"Hofer Bau GmbH", "Müller AG"}:
            return mapping.setdefault(text, f"<COMPANY_{len(mapping) + 1:03d}>")
        return text

    result = XlsxHandler().process(inp, out, transform=transform)
    assert result.segments_processed > 0

    from openpyxl import load_workbook

    wb = load_workbook(str(out))
    ws = wb.active
    # Spalte A — identische Werte → identische Platzhalter
    assert ws["A2"].value == ws["A3"].value
    assert ws["A2"].value.startswith("<COMPANY_")
    # Numerische Saldo-Werte unverändert
    assert ws["B2"].value == 1000
    # Formel-Struktur erhalten, String-Literal anonymisiert
    formula_cell = ws["C2"]
    assert formula_cell.value.startswith("=SUMIF")
    assert "Hofer Bau GmbH" not in str(formula_cell.value)
    assert "<COMPANY_" in str(formula_cell.value)
    # Reine Formel ohne String-Literal bleibt unverändert
    assert ws["C3"].value == "=SUM(B2:B4)"


def test_xlsx_empty_formula_string_is_left_alone(tmp_path: Path) -> None:
    from openpyxl import Workbook, load_workbook

    inp = tmp_path / "empty_lit.xlsx"
    out = tmp_path / "empty_lit.anon.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = '=IF(B1="","leer","gefüllt")'
    wb.save(str(inp))

    calls: list[str] = []

    def transform(text: str) -> str:
        calls.append(text)
        return text.upper()

    XlsxHandler().process(inp, out, transform=transform)

    wb_out = load_workbook(str(out))
    formula = wb_out.active["A1"].value
    # Leeres String-Literal "" bleibt unangetastet, andere werden transformiert.
    assert '""' in formula
    assert "LEER" in formula and "GEFÜLLT" in formula


def test_xlsx_default_output_and_supports() -> None:
    h = XlsxHandler()
    assert h.supports(Path("a.xlsx"))
    assert not h.supports(Path("a.csv"))
    assert h.default_output_path(Path("a.xlsx"), "anon") == Path("a.anon.xlsx")

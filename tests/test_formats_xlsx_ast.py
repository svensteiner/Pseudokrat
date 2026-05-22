"""XLSX-Tokenizer-Tests — verifizieren, dass das Formel-Parsen sich nicht mehr
mit Regex-Heuristiken begnügt, sondern den echten Excel-Token-Strom benutzt.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from pseudokrat.formats.xlsx_handler import XlsxHandler, _transform_formula


def _upper(text: str) -> str:
    return text.upper()


def test_transform_keeps_cell_refs_untouched() -> None:
    formula = '=SUMIF(A:A,"hofer bau gmbh",B:B)'
    out = _transform_formula(formula, _upper)
    assert out == '=SUMIF(A:A,"HOFER BAU GMBH",B:B)'


def test_transform_keeps_sheet_qualified_refs_untouched() -> None:
    formula = "=SUM('Mandant Hofer'!A1:A10)"
    out = _transform_formula(formula, _upper)
    # Sheet-Name in Single-Quotes ist KEIN String-Literal, sondern eine
    # Reference. Darf NICHT transformiert werden.
    assert out == "=SUM('Mandant Hofer'!A1:A10)"


def test_transform_handles_escaped_quotes_inside_string() -> None:
    """In Excel werden doppelte Quotes innerhalb von Strings escaped."""
    formula = '=CONCATENATE("Bob said ""hi""")'
    out = _transform_formula(formula, _upper)
    # Sowohl der innere String-Inhalt als auch das Quote-Escaping müssen
    # korrekt durchgereicht werden.
    assert out.startswith("=CONCATENATE(")
    assert 'BOB SAID ""HI""' in out


def test_transform_keeps_function_names_untouched() -> None:
    formula = '=IF(A1="Hofer Bau GmbH","ja","nein")'
    out = _transform_formula(formula, _upper)
    assert "IF(" in out
    assert "A1=" in out
    assert '"HOFER BAU GMBH"' in out
    assert '"JA"' in out
    assert '"NEIN"' in out


def test_transform_keeps_numbers_untouched() -> None:
    formula = '=A1*1.5+SUM(B1:B100,"Mandant",4711)'
    out = _transform_formula(formula, _upper)
    assert "1.5" in out
    assert "4711" in out
    assert '"MANDANT"' in out


def test_transform_handles_named_range_argument() -> None:
    """Named-Ranges in Formel-Argumenten dürfen NICHT angepackt werden."""
    formula = '=SUMIF(MWST_SATZ,">0","tax")'
    out = _transform_formula(formula, _upper)
    assert "MWST_SATZ" in out
    assert '">0"' in out  # ">0" ist ein String-Literal-Operand, aber idempotent
    assert '"TAX"' in out


def test_handler_processes_named_ranges_in_workbook(tmp_path: Path) -> None:
    """Workbook-Level-Named-Range mit String-Literal in der Formel."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Hofer Bau GmbH"
    ws["B1"] = 100

    # Named Range mit String-Literal:
    from openpyxl.workbook.defined_name import DefinedName

    dn = DefinedName(name="MandantenFilter", attr_text='"Hofer Bau GmbH"')
    wb.defined_names["MandantenFilter"] = dn

    input_path = tmp_path / "in.xlsx"
    output_path = tmp_path / "out.xlsx"
    wb.save(input_path)

    handler = XlsxHandler()
    result = handler.process(input_path, output_path, transform=lambda s: f"<{s}>")
    assert result.output_path == output_path

    from openpyxl import load_workbook

    wb2 = load_workbook(output_path)
    a1 = wb2.active["A1"].value
    assert a1 == "<Hofer Bau GmbH>"
    dn2 = wb2.defined_names["MandantenFilter"]
    attr = getattr(dn2, "attr_text", None) or getattr(dn2, "value", None)
    assert attr == '"<Hofer Bau GmbH>"'


def test_handler_processes_cross_sheet_formula(tmp_path: Path) -> None:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Saldo"
    ws1["A1"] = "Hofer Bau GmbH"
    ws1["B1"] = 250

    ws2 = wb.create_sheet("Pivot")
    ws2["A1"] = '=SUMIF(Saldo!A:A,"Hofer Bau GmbH",Saldo!B:B)'
    ws2["A1"].data_type = "f"

    input_path = tmp_path / "in.xlsx"
    output_path = tmp_path / "out.xlsx"
    wb.save(input_path)

    handler = XlsxHandler()
    handler.process(input_path, output_path, transform=lambda s: f"<{s}>")

    from openpyxl import load_workbook

    wb2 = load_workbook(output_path)
    formula = wb2["Pivot"]["A1"].value
    # Sheet-Referenz „Saldo!" muss intakt sein:
    assert "Saldo!A:A" in formula
    assert "Saldo!B:B" in formula
    # String-Literal muss transformiert sein:
    assert '"<Hofer Bau GmbH>"' in formula


@pytest.mark.parametrize(
    ("formula", "expected_inner_count"),
    [
        ('=A1+"foo"+"bar"', 2),
        ('=IF(A1>0,"j","n")', 2),
        ('=CONCAT("a","b","c")', 3),
        ("=A1+A2", 0),
    ],
)
def test_transform_counts_string_literals_correctly(
    formula: str, expected_inner_count: int
) -> None:
    seen: list[str] = []

    def collect(value: str) -> str:
        seen.append(value)
        return value.upper()

    _transform_formula(formula, collect)
    assert len(seen) == expected_inner_count

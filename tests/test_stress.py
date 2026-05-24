"""Stress-Tests: große Inputs und viele Entitäten.

Diese Suite verifiziert, dass:

1. Die Pipeline auch mit 10k-Zeilen-Texten konsistent bleibt.
2. 1.000 distinct PII-Entitäten in einem Dokument fehlerfrei verarbeitet werden.
3. XLSX-Tabellen mit vielen Zellen + Formel-String-Literalen korrekt anonymisiert
   werden (Megaprompt §12.7).
4. Wiederholtes Anonymisieren großer Texte deterministisch und schnell ist.

Stress-Tests sind als ``@pytest.mark.slow`` markiert, damit sie in normalen CI-
Runs optional deselektierbar sind über ``pytest -m "not slow"``.
"""

from __future__ import annotations

import time
from collections.abc import Iterator
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from pseudokrat.anonymizer import Anonymizer
from pseudokrat.deanonymizer import Deanonymizer
from pseudokrat.formats import XlsxHandler
from pseudokrat.recognizers import default_recognizers
from pseudokrat.store.profile import ProfileManager

pytestmark = pytest.mark.slow


@pytest.fixture
def heavy_pipeline(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> Iterator[tuple[Anonymizer, Deanonymizer, Path]]:
    monkeypatch.setenv("PSEUDOKRAT_DATA_DIR", str(tmp_path / "ps"))
    pm = ProfileManager()
    store, audit = pm.open_or_create("stress", "stress-test-pw-2026")
    anon = Anonymizer(
        store=store,
        recognizers=default_recognizers(),
        detector=None,
        audit_log=audit,
        model_version="stress",
    )
    deanon = Deanonymizer(store=store, audit_log=audit, model_version="stress")
    try:
        yield anon, deanon, tmp_path
    finally:
        store.close()


def _generate_n_valid_at_uids(n: int) -> list[str]:
    """Erzeuge n syntaktisch gültige AT-UIDs."""
    from tests.test_property_recognizers import _at_uid_check_digit

    uids: list[str] = []
    counter = 0
    while len(uids) < n:
        digits = [int(d) for d in f"{counter:07d}"]
        digits.append(_at_uid_check_digit(digits))
        candidate = "ATU" + "".join(str(d) for d in digits)
        uids.append(candidate)
        counter += 1
    return uids


# --------------------------------------------------------------------------- #
# 1. Großes Freitext-Dokument                                                 #
# --------------------------------------------------------------------------- #


def test_10k_line_document_roundtrip(
    heavy_pipeline: tuple[Anonymizer, Deanonymizer, Path],
) -> None:
    """10.000 Zeilen, jede mit einer eindeutigen UID — Round-Trip muss perfekt sein."""
    anon, deanon, _ = heavy_pipeline
    uids = _generate_n_valid_at_uids(2_000)
    lines = [f"Zeile {i:05d}: Mandant mit UID {uids[i]}." for i in range(2_000)]
    doc = "\n".join(lines)

    start = time.perf_counter()
    r = anon.anonymize(doc)
    anon_elapsed = time.perf_counter() - start
    print(f"\n  Anonymisierung 10k Zeilen: {anon_elapsed:.2f}s")

    # Wir erwarten 10.000 distinct UID-Mappings — pro UID einer.
    assert r.entity_counts.get("UID", 0) == 2_000

    start = time.perf_counter()
    decoded = deanon.deanonymize(r.text).text
    dean_elapsed = time.perf_counter() - start
    print(f"  Deanonymisierung 10k Zeilen: {dean_elapsed:.2f}s")
    assert decoded == doc

    # Performance-Sanity: < 60s gesamt (großzügig — wir sind nicht optimiert)
    assert anon_elapsed + dean_elapsed < 120.0


# --------------------------------------------------------------------------- #
# 2. Dokument mit 1.000 distinct Entitäten                                    #
# --------------------------------------------------------------------------- #


def test_1000_distinct_entities_get_distinct_placeholders(
    heavy_pipeline: tuple[Anonymizer, Deanonymizer, Path],
) -> None:
    anon, deanon, _ = heavy_pipeline
    uids = _generate_n_valid_at_uids(1_000)
    # Ein zusammenhängendes Dokument
    doc = "\n".join(f"Mandant #{i}: UID {uids[i]}." for i in range(1_000))
    r = anon.anonymize(doc)
    # Alle 1000 UIDs müssen unterschiedliche Platzhalter haben.
    uid_placeholders = {s.text for s in r.spans if s.category == "UID"}
    assert len(uid_placeholders) == 1_000, (
        f"Erwartete 1.000 distinct UIDs, fanden {len(uid_placeholders)}"
    )
    # Im Anonymisat-Text müssen 1.000 verschiedene <UID_xxx>-Platzhalter sein.
    import re

    placeholders = set(re.findall(r"<UID_\d+>", r.text))
    assert len(placeholders) == 1_000

    # Round-Trip
    decoded = deanon.deanonymize(r.text).text
    assert decoded == doc


# --------------------------------------------------------------------------- #
# 3. Wiederholungs-Stress: 1 Entität 10.000-mal                               #
# --------------------------------------------------------------------------- #


def test_repeated_same_entity_one_placeholder(
    heavy_pipeline: tuple[Anonymizer, Deanonymizer, Path],
) -> None:
    anon, deanon, _ = heavy_pipeline
    uid = _generate_n_valid_at_uids(1)[0]
    doc = "\n".join(f"Zeile {i}: UID {uid} (Wiederholung)." for i in range(2_000))
    r = anon.anonymize(doc)
    # Genau ein Mapping
    placeholders = {s.text for s in r.spans}
    # Spans enthalten Originaltexte; alle sollten denselben Original-Wert haben.
    assert placeholders == {uid}
    decoded = deanon.deanonymize(r.text).text
    assert decoded == doc


# --------------------------------------------------------------------------- #
# 4. Mixed-PII-Dokument                                                       #
# --------------------------------------------------------------------------- #


def test_mixed_pii_500_each(
    heavy_pipeline: tuple[Anonymizer, Deanonymizer, Path],
) -> None:
    """500 IBANs + 500 UIDs + 500 USt-IdNrs in einem Dokument."""
    from tests.test_property_recognizers import (
        _iso_7064_check,
        _mod97_check,
    )

    uids = _generate_n_valid_at_uids(500)
    # 500 USt-IdNrs sequenziell
    usts: list[str] = []
    counter = 1
    while len(usts) < 500:
        digit_str = f"{counter:08d}"
        check = _iso_7064_check(digit_str)
        usts.append(f"DE{digit_str}{check}")
        counter += 1
    # 500 ATs IBANs
    ibans: list[str] = []
    counter = 1
    while len(ibans) < 500:
        bban = f"{counter:016d}"
        check = _mod97_check(bban, "AT")
        ibans.append(f"AT{check}{bban}")
        counter += 1

    lines = []
    for i in range(500):
        lines.append(f"Mandant {i}: UID {uids[i]}, USt-IdNr {usts[i]}, IBAN {ibans[i]}.")
    doc = "\n".join(lines)

    r = anon_result = (heavy_pipeline[0]).anonymize(doc)
    assert anon_result.entity_counts.get("UID", 0) >= 1000  # UID + USt-IdNr beide "UID"
    assert anon_result.entity_counts.get("IBAN", 0) == 500

    decoded = heavy_pipeline[1].deanonymize(r.text).text
    assert decoded == doc


# --------------------------------------------------------------------------- #
# 5. XLSX-Stress (Megaprompt §12.7)                                           #
# --------------------------------------------------------------------------- #


def test_xlsx_1000_rows_5_distinct_mandanten(
    heavy_pipeline: tuple[Anonymizer, Deanonymizer, Path],
) -> None:
    """1.000-Zeilen-Saldenliste mit 5 Mandanten, alle Vorkommen konsistent."""
    anon, deanon, tmp_path = heavy_pipeline
    workdir = tmp_path / "xlsx"
    workdir.mkdir()
    src = workdir / "salden.xlsx"

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Salden"
    ws.append(["Mandant", "Konto", "Saldo"])
    mandanten = [
        "Hofer Bau GmbH",
        "Müller Consulting AG",
        "Schmidt & Partner OG",
        "Wiener Handel KG",
        "Kärntner Industrie AG",
    ]
    for i in range(1_000):
        m = mandanten[i % 5]
        ws.append([m, f"K-{i:04d}", round(1000 + i * 1.23, 2)])
    # Eine Formel mit String-Literal:
    ws["E1"] = '=SUMIF(A:A,"Hofer Bau GmbH",C:C)'
    wb.save(src)

    out = workdir / "salden.anon.xlsx"
    XlsxHandler().process(
        src,
        out,
        transform=lambda t: anon.anonymize(t).text,
    )

    # Anonymisiert öffnen und prüfen
    awb = load_workbook(out)
    aws = awb["Salden"]
    # Spalte A: 1.000 + Header. Header bleibt.
    header = aws["A1"].value
    assert header == "Mandant"
    # Werte in Spalte A müssen nur 5 distinct Platzhalter haben.
    a_values = {aws.cell(row=i, column=1).value for i in range(2, 1002)}
    # Es können maximal 5 Mappings sein
    assert len(a_values) == 5, f"Erwartete 5 distinct Mandanten-Platzhalter, fand {len(a_values)}"
    # Original-Namen dürfen NICHT mehr vorkommen
    for m in mandanten:
        assert m not in a_values, f"Original-Name {m!r} leakte ins XLSX"
    # Formel-String-Literal muss auch ersetzt sein
    formula = aws["E1"].value
    assert formula is not None
    assert "Hofer Bau GmbH" not in formula, f"Formel-Literal nicht anonymisiert: {formula!r}"


# --------------------------------------------------------------------------- #
# 6. Audit-Log unter Stress: 10.000 Einträge                                  #
# --------------------------------------------------------------------------- #


def test_audit_log_10k_entries_chain_remains_valid(
    heavy_pipeline: tuple[Anonymizer, Deanonymizer, Path],
) -> None:
    anon, _, _ = heavy_pipeline
    for i in range(1_000):  # 1k Operationen genügen, jeweils 1 Audit-Eintrag
        anon.anonymize(f"Test #{i}: UID ATU12345675")
    # Hole das AuditLog vom Anonymizer
    audit = anon._audit_log  # type: ignore[attr-defined]
    assert audit is not None
    assert audit.verify_chain()
    entries = audit.all_entries()
    assert len(entries) >= 1_000

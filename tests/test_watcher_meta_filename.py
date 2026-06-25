"""Tests fuer Dateiname-Anonymisierung und Office-Metadaten-Entfernung."""

from __future__ import annotations

import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

from pseudokrat import watcher


class _FakeAnonymizer:
    """Minimaler Anonymizer-Ersatz: ersetzt nur 'Hankook'."""

    def anonymize(self, text: str):  # noqa: ANN001 - Test-Stub
        class _Result:
            pass

        r = _Result()
        r.text = text.replace("Hankook", "<COMPANY_001>")
        return r


class TestSafeFilename:
    def test_brackets_removed_and_whitespace_collapsed(self) -> None:
        stem = watcher.safe_anonymized_stem("Hankook Bilanz 2025", _FakeAnonymizer())
        assert "<" not in stem and ">" not in stem
        assert "Hankook" not in stem
        assert stem == "COMPANY_001_Bilanz_2025"

    def test_invalid_windows_chars_neutralised(self) -> None:
        stem = watcher.safe_anonymized_stem('A/B:C?D', _FakeAnonymizer())
        for bad in '<>:"/\\|?*':
            assert bad not in stem

    def test_empty_fallback(self) -> None:
        assert watcher.safe_anonymized_stem("   ", _FakeAnonymizer()) == "dokument"


class TestStripOfficeMetadata:
    def test_xlsx_properties_removed(self, tmp_path: Path) -> None:
        path = tmp_path / "buch.xlsx"
        wb = Workbook()
        wb.properties.creator = "Hankook GmbH"
        wb.properties.lastModifiedBy = "Max Mustermann"
        wb.properties.title = "Bilanz Hankook"
        wb.active["A1"] = "x"
        wb.save(path)

        watcher.strip_office_metadata(path)

        wb2 = load_workbook(path)
        # Entscheidend ist, dass kein eingebetteter Mandanten-/Autorenname
        # ueberlebt — NICHT, dass jedes Feld leer ist: openpyxl setzt bei
        # fehlendem <dc:creator> den generischen Default "openpyxl" (kein PII).
        # Der Roh-XML-Nachweis liegt in test_core_xml_has_no_author.
        leftover = " ".join(
            str(v)
            for v in (
                wb2.properties.creator,
                wb2.properties.lastModifiedBy,
                wb2.properties.title,
                wb2.properties.subject,
                wb2.properties.description,
                wb2.properties.keywords,
            )
            if v
        )
        assert "Hankook" not in leftover
        assert "Max Mustermann" not in leftover
        assert not wb2.properties.lastModifiedBy
        assert not wb2.properties.title
        # Datei bleibt eine gueltige, lesbare XLSX.
        assert wb2.active["A1"].value == "x"

    def test_non_office_is_noop(self, tmp_path: Path) -> None:
        p = tmp_path / "a.txt"
        p.write_text("hallo", encoding="utf-8")
        watcher.strip_office_metadata(p)  # darf nicht scheitern
        assert p.read_text(encoding="utf-8") == "hallo"

    def test_core_xml_has_no_author(self, tmp_path: Path) -> None:
        path = tmp_path / "doc.xlsx"
        wb = Workbook()
        wb.properties.creator = "Geheim AG"
        wb.save(path)
        watcher.strip_office_metadata(path)
        with zipfile.ZipFile(path) as z:
            core = z.read("docProps/core.xml").decode("utf-8")
        assert "Geheim AG" not in core

"""Erzeugt eine grosse synthetische Test-XLSX mit oesterreichischen PII-Daten.

Alle IBAN/UID/SVNR werden gegen Pseudokrats EIGENE Erkenner validiert, damit
die Testdatei garantiert greift (Pruefziffern korrekt). KEINE echten
Mandantendaten — alles synthetisch (Seed 42, reproduzierbar).

Aufruf (im aktivierten venv mit installiertem pseudokrat):

    python tools/gen_testdaten.py [AUSGABE.xlsx]

Standard-Ausgabe: ./Testdaten_AT_gross.xlsx im aktuellen Verzeichnis.

Tabellenblaetter: Debitoren, Kreditoren, Mitarbeiter (mit SVNR), Stammdaten,
Buchungen (gross, mit SUMIF-Formel zum Test der Formel-Konsistenz).
"""

from __future__ import annotations

import os
import random
import sys
from pathlib import Path

os.environ.setdefault("PSEUDOKRAT_DISABLE_ML", "1")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from pseudokrat.anonymizer import Anonymizer
from pseudokrat.recognizers import recognizers_for_store
from pseudokrat.store.profile import ProfileManager

OUT = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("Testdaten_AT_gross.xlsx")

rnd = random.Random(42)  # reproduzierbar

manager = ProfileManager()
store, audit = manager.open_or_create_simple("GenTmp")  # Wegwerf-Profil
with store:
    anon = Anonymizer(
        store=store, recognizers=recognizers_for_store(store),
        detector=None, audit_log=None, model_version="disabled",
    )

    def categories(text: str) -> set[str]:
        return {s.category for s in anon.detect(text)}

    # ---- IBAN (AT, MOD-97) ----
    def iban_check(country: str, bban: str) -> str:
        rearr = bban + country + "00"
        num = "".join(str(ord(c) - 55) if c.isalpha() else c for c in rearr)
        return f"{98 - (int(num) % 97):02d}"

    def gen_iban() -> str:
        for _ in range(50):
            bban = "".join(rnd.choice("0123456789") for _ in range(16))
            iban = "AT" + iban_check("AT", bban) + bban
            if "IBAN" in categories(iban):
                return iban
        raise RuntimeError("keine gueltige IBAN erzeugt")

    # ---- UID / SVNR: Pruefziffer per Brute-Force gegen den Erkenner ----
    def gen_uid() -> str:
        for _ in range(60):
            base = "".join(rnd.choice("0123456789") for _ in range(7))
            for chk in "0123456789":
                cand = "ATU" + base + chk
                if "UID" in categories(cand):
                    return cand
        raise RuntimeError("keine gueltige UID erzeugt")

    def gen_svnr() -> str:
        for _ in range(200):
            run = "".join(rnd.choice("0123456789") for _ in range(3))
            day = f"{rnd.randint(1, 28):02d}"
            mon = f"{rnd.randint(1, 12):02d}"
            yr = f"{rnd.randint(45, 99):02d}"
            for chk in "0123456789":
                cand = f"{run}{chk} {day}{mon}{yr}"
                if "SVNR" in categories(cand):
                    return cand
        raise RuntimeError("keine gueltige SVNR erzeugt")

    VORNAMEN = ["Anna", "Lukas", "Maria", "Florian", "Julia", "Stefan", "Sandra",
                "Markus", "Birgit", "Thomas", "Eva", "Christoph", "Petra", "Andreas"]
    NACHNAMEN = ["Huber", "Gruber", "Wagner", "Maier", "Steiner", "Moser", "Bauer",
                 "Pichler", "Hofer", "Leitner", "Berger", "Fuchs", "Eder", "Lang"]
    RECHTSFORM = ["GmbH", "AG", "KG", "OG", "e.U.", "GmbH & Co KG"]
    FIRMENWORT = ["Alpen", "Donau", "Tirol", "Wien", "Styria", "Berg", "Sonnen",
                  "Nord", "Central", "Prime", "Edel", "Austro", "Salzach", "Wald"]
    BRANCHE = ["Bau", "Handel", "Consulting", "Immobilien", "Technik", "Gastro",
               "Logistik", "Pharma", "Energie", "Druck"]
    ORTE = [("Wien", "1010"), ("Graz", "8010"), ("Linz", "4020"), ("Salzburg", "5020"),
            ("Innsbruck", "6020"), ("Klagenfurt", "9020"), ("Wels", "4600")]
    STRASSEN = ["Hauptstrasse", "Bahnhofstrasse", "Ringstrasse", "Mariahilfer Strasse",
                "Lindengasse", "Schulgasse", "Feldweg", "Industriestrasse"]

    def person() -> str:
        return f"{rnd.choice(VORNAMEN)} {rnd.choice(NACHNAMEN)}"

    def firma() -> str:
        return f"{rnd.choice(FIRMENWORT)} {rnd.choice(BRANCHE)} {rnd.choice(RECHTSFORM)}"

    def email(name: str, fa: str) -> str:
        n = (name.lower().replace(" ", ".").replace("ä", "ae")
             .replace("ö", "oe").replace("ü", "ue"))
        return f"{n}@{fa.split()[0].lower()}.at"

    def telefon() -> str:
        return f"+43 {rnd.randint(1, 7)}{rnd.randint(10, 99)} {rnd.randint(100000, 999999)}"

    def adresse() -> str:
        ort, plz = rnd.choice(ORTE)
        return f"{rnd.choice(STRASSEN)} {rnd.randint(1, 180)}, {plz} {ort}"

    wb = Workbook()
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="305496")

    def write_header(ws, headers):
        ws.append(headers)
        for c in ws[1]:
            c.font = head_font
            c.fill = head_fill

    print("Generiere gueltige IBAN/UID/SVNR ...", flush=True)
    ibans = [gen_iban() for _ in range(60)]
    uids = [gen_uid() for _ in range(40)]
    svnrs = [gen_svnr() for _ in range(90)]
    firmen = [firma() for _ in range(40)]
    print(f"  IBAN={len(ibans)} UID={len(uids)} SVNR={len(svnrs)} Firmen={len(firmen)}",
          flush=True)

    ws = wb.active
    ws.title = "Debitoren"
    write_header(ws, ["KdNr", "Firma", "Ansprechpartner", "UID", "IBAN", "E-Mail",
                      "Telefon", "Adresse", "Offener Betrag"])
    for i in range(60):
        fa = rnd.choice(firmen)
        p = person()
        ws.append([f"D{1000 + i}", fa, p, rnd.choice(uids), rnd.choice(ibans),
                   email(p, fa), telefon(), adresse(), round(rnd.uniform(500, 90000), 2)])

    ws = wb.create_sheet("Kreditoren")
    write_header(ws, ["LiefNr", "Firma", "Ansprechpartner", "UID", "IBAN", "E-Mail",
                      "Telefon", "Adresse", "Verbindlichkeit"])
    for i in range(55):
        fa = rnd.choice(firmen)
        p = person()
        ws.append([f"K{2000 + i}", fa, p, rnd.choice(uids), rnd.choice(ibans),
                   email(p, fa), telefon(), adresse(), round(rnd.uniform(200, 60000), 2)])

    ws = wb.create_sheet("Mitarbeiter")
    write_header(ws, ["PersNr", "Name", "SVNR", "Gehalts-IBAN", "E-Mail",
                      "Telefon", "Adresse", "Bruttogehalt"])
    for i in range(90):
        p = person()
        ws.append([f"P{500 + i}", p, rnd.choice(svnrs), rnd.choice(ibans),
                   email(p, "lohnbuero"), telefon(), adresse(),
                   round(rnd.uniform(2200, 9500), 2)])

    ws = wb.create_sheet("Stammdaten")
    write_header(ws, ["Feld", "Wert"])
    haupt = firmen[0]
    for row in [
        ["Mandant", haupt],
        ["UID", uids[0]],
        ["Bank-IBAN", ibans[0]],
        ["Geschaeftsfuehrer", person()],
        ["Kontakt E-Mail", email("office", haupt)],
        ["Telefon", telefon()],
        ["Sitz", adresse()],
        ["Website", f"https://www.{haupt.split()[0].lower()}.at"],
    ]:
        ws.append(row)

    ws = wb.create_sheet("Buchungen")
    write_header(ws, ["BuchNr", "Datum", "Gegenkonto (Firma)", "IBAN", "Text",
                      "Soll", "Haben"])
    N = 4000
    for i in range(N):
        fa = rnd.choice(firmen)
        soll = round(rnd.uniform(0, 25000), 2) if rnd.random() < 0.5 else 0
        haben = 0 if soll else round(rnd.uniform(0, 25000), 2)
        ws.append([i + 1, f"2026-{rnd.randint(1, 12):02d}-{rnd.randint(1, 28):02d}",
                   fa, rnd.choice(ibans), f"Rechnung {fa}", soll, haben])
    sumrow = N + 3
    ws.cell(row=sumrow, column=3, value="Summe Soll fuer:")
    ws.cell(row=sumrow, column=4, value=firmen[0])
    ws.cell(row=sumrow, column=6, value=f'=SUMIF(C2:C{N + 1},"{firmen[0]}",F2:F{N + 1})')

    wb.save(OUT)
    print(f"Gespeichert: {OUT.resolve()} ({OUT.stat().st_size} Bytes)", flush=True)

# Wegwerf-Profil entfernen, damit es die Profilliste nicht zumuellt.
try:
    p = manager.profile_path("GenTmp")
    for f in (p, p.with_suffix(p.suffix + ".keyring")):
        if f.exists():
            f.unlink()
except OSError as exc:
    print("Hinweis: Wegwerf-Profil konnte nicht entfernt werden:", exc)
